import os
import json
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fastapi import HTTPException
import firebase_admin
from firebase_admin import credentials, storage

class ExcelExportServiceV2:
    """
    Servicio para exportar reportes de Excel usando plantillas con openpyxl
    para preservar estilos y formato profesional.
    
    Soporta múltiples fuentes para la plantilla:
    1. Cloud Storage (recomendado para producción)
    2. Sistema de archivos local (fallback para desarrollo)
    """
    
    def __init__(self):
        # Configuración de Cloud Storage
        self.bucket_name = "west-reportes-conduccion.appspot.com"
        self.template_blob_name = "excel-templates/referenciaReporte.xlsx"
        
        # Ruta local fallback (para desarrollo)
        self.local_template_path = os.path.join(os.path.dirname(__file__), '..', '..', 'referenciaReporte.xlsx')
        
        # Mapeo de nombres de alarmas
        self.alarm_name_mapping = {
            'cinturon': 'Cinturón de seguridad',
            'distraido': 'Conductor distraído',
            'cruce': 'Cruce de carril',
            'distancia': 'Distancia de seguridad',
            'fatiga': 'Fatiga',
            'frenada': 'Frenada brusca',
            'stop': 'Infracción de señal de stop',
            'telefono': 'Teléfono móvil',
            'boton': 'Botón de Alerta',
            'video': 'Video Solicitado',
        }
        
        # Inicializar cliente de Storage
        self._initialize_storage()
    
    def _initialize_storage(self):
        """Inicializa el cliente de Firebase Storage"""
        try:
            if not firebase_admin._apps:
                # Ya está inicializado en firebase.py, pero por si acaso
                cred = credentials.ApplicationDefault()
                firebase_admin.initialize_app(cred)
            
            # CORREGIDO: Usar el método correcto para obtener el bucket directamente
            self.bucket = storage.bucket(self.bucket_name)
            self.storage_client = storage
            
        except Exception as e:
            print(f"Error al inicializar Storage: {e}")
            self.storage_client = None
            self.bucket = None
    
    def _download_template_from_storage(self) -> Optional[str]:
        """
        Descarga la plantilla desde Cloud Storage
        
        Returns:
            str: Ruta al archivo temporal descargado, o None si falla
        """
        if not self.bucket:
            return None
        
        try:
            # Crear archivo temporal
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            
            # Descargar el blob
            blob = self.bucket.blob(self.template_blob_name)
            blob.download_to_filename(temp_file.name)
            
            print(f"Plantilla descargada desde Storage: {temp_file.name}")
            return temp_file.name
            
        except Exception as e:
            print(f"Error al descargar plantilla desde Storage: {e}")
            # Limpiar archivo temporal si existe
            if 'temp_file' in locals():
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
            return None
    
    def _get_template_path(self) -> str:
        """
        Obtiene la ruta a la plantilla de Excel, intentando primero Cloud Storage
        
        Returns:
            str: Ruta al archivo de plantilla
            
        Raises:
            HTTPException: Si no se encuentra la plantilla en ninguna fuente
        """
        # Intentar descargar desde Cloud Storage primero
        storage_path = self._download_template_from_storage()
        if storage_path:
            return storage_path
        
        # Fallback a sistema de archivos local
        if os.path.exists(self.local_template_path):
            print(f"Usando plantilla local: {self.local_template_path}")
            return self.local_template_path
        
        # Si ninguna opción funciona, lanzar error
        raise HTTPException(
            status_code=500,
            detail="Plantilla de Excel no encontrada en Cloud Storage ni en sistema local"
        )
    
    def _cleanup_temp_file(self, file_path: str):
        """Limpia archivos temporales"""
        try:
            if file_path and file_path != self.local_template_path:
                os.unlink(file_path)
                print(f"Archivo temporal limpiado: {file_path}")
        except Exception as e:
            print(f"Error al limpiar archivo temporal {file_path}: {e}")
    
    def _copy_cell_style(self, source_cell, target_cell):
        """
        Copia el estilo de una celda a otra
        """
        if source_cell:
            # Copiar fuente
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    underline=source_cell.font.underline,
                    strikethrough=source_cell.font.strikethrough,
                    color=source_cell.font.color
                )
            
            # Copiar relleno
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # Copiar alineación
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # Copiar bordes
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            
            # Copiar formato de número
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
    
    def _get_cell_style(self, cell):
        """
        Obtiene el estilo de una celda como un diccionario
        """
        if not cell:
            return None
        
        style = {}
        
        if cell.font:
            style['font'] = {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'underline': cell.font.underline,
                'strikethrough': cell.font.strikethrough,
                'color': cell.font.color
            }
        
        if cell.fill:
            style['fill'] = {
                'fill_type': cell.fill.fill_type,
                'start_color': cell.fill.start_color,
                'end_color': cell.fill.end_color
            }
        
        if cell.alignment:
            style['alignment'] = {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'wrap_text': cell.alignment.wrap_text,
                'shrink_to_fit': cell.alignment.shrink_to_fit,
                'indent': cell.alignment.indent
            }
        
        if cell.border:
            style['border'] = {
                'left': cell.border.left,
                'right': cell.border.right,
                'top': cell.border.top,
                'bottom': cell.border.bottom
            }
        
        if cell.number_format:
            style['number_format'] = cell.number_format
        
        return style
    
    def _apply_cell_style(self, cell, style):
        """
        Aplica un estilo a una celda
        """
        if not style or not cell:
            return
        
        if 'font' in style:
            font_style = style['font']
            cell.font = Font(
                name=font_style.get('name'),
                size=font_style.get('size'),
                bold=font_style.get('bold'),
                italic=font_style.get('italic'),
                underline=font_style.get('underline'),
                strikethrough=font_style.get('strikethrough'),
                color=font_style.get('color')
            )
        
        if 'fill' in style:
            fill_style = style['fill']
            cell.fill = PatternFill(
                fill_type=fill_style.get('fill_type'),
                start_color=fill_style.get('start_color'),
                end_color=fill_style.get('end_color')
            )
        
        if 'alignment' in style:
            alignment_style = style['alignment']
            cell.alignment = Alignment(
                horizontal=alignment_style.get('horizontal'),
                vertical=alignment_style.get('vertical'),
                wrap_text=alignment_style.get('wrap_text'),
                shrink_to_fit=alignment_style.get('shrink_to_fit'),
                indent=alignment_style.get('indent')
            )
        
        if 'border' in style:
            border_style = style['border']
            cell.border = Border(
                left=border_style.get('left'),
                right=border_style.get('right'),
                top=border_style.get('top'),
                bottom=border_style.get('bottom')
            )
        
        if 'number_format' in style:
            cell.number_format = style['number_format']
    
    def _update_cell_preserve_style(self, sheet, cell_ref: str, value: Any, cell_type: str = 's'):
        """
        Actualiza una celda preservando su estilo existente
        """
        try:
            # Si la celda no existe, la creamos
            if cell_ref not in sheet:
                sheet[cell_ref] = ''
            
            # Obtener la celda y guardar su estilo actual
            cell = sheet[cell_ref]
            current_style = self._get_cell_style(cell)
            
            # Actualizar el valor
            cell.value = value
            
            # Forzar el tipo de celda
            cell.data_type = cell_type
            
            # Restaurar el estilo original
            if current_style:
                self._apply_cell_style(cell, current_style)
            
        except Exception as e:
            print(f"Error al actualizar celda {cell_ref}: {e}")
    
    def _format_timestamp(self, timestamp: str) -> str:
        """
        Formatea un timestamp para mostrar en Excel, ajustando a la zona horaria America/Santiago (UTC-3)
        """
        try:
            # Zona horaria de Santiago
            santiago_tz = ZoneInfo('America/Santiago')
            
            # Si es un timestamp en formato Unix
            if timestamp.isdigit():
                # Convertir timestamp UTC a datetime en Santiago
                dt_utc = datetime.fromtimestamp(int(timestamp), tz=ZoneInfo('UTC'))
                dt_santiago = dt_utc.astimezone(santiago_tz)
                return dt_santiago.strftime('%d/%m/%Y %H:%M:%S')
            
            # Si ya es una fecha en formato string, intentar formatearla
            try:
                # Reemplazar 'Z' con UTC y parsear
                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                
                # Si no tiene información de zona horaria, asumir que es UTC
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=ZoneInfo('UTC'))
                
                # Convertir a zona horaria de Santiago
                dt_santiago = dt.astimezone(santiago_tz)
                return dt_santiago.strftime('%d/%m/%Y %H:%M:%S')
            except:
                return timestamp
                
        except Exception as e:
            print(f"Error al formatear timestamp {timestamp}: {e}")
            return timestamp
    
    def export_report_to_excel(self, data: Dict[str, Any]) -> bytes:
        """
        Genera un archivo Excel exportado desde los datos del reporte
        
        Args:
            data: Diccionario con los datos del reporte
            
        Returns:
            bytes: Contenido del archivo Excel generado
        """
        template_path = None
        try:
            # Obtener la plantilla (desde Storage o local)
            template_path = self._get_template_path()
            
            # Cargar la plantilla
            workbook = load_workbook(template_path)
            
            # Extraer datos del reporte
            current_report = data.get('current_report', {})
            filtered_events = data.get('filtered_events', [])
            selected_company = data.get('selected_company', 'N/A')
            
            # --- ACTUALIZAR HOJA DE RESUMEN ---
            summary_sheet = workbook.get_sheet_by_name('Resumen')
            if summary_sheet:
                # Actualizar información del reporte
                self._update_cell_preserve_style(summary_sheet, 'B2', selected_company, 's')
                self._update_cell_preserve_style(summary_sheet, 'B3', current_report.get('vehicle_plate', 'N/A'), 's')
                self._update_cell_preserve_style(summary_sheet, 'B4', current_report.get('file_name', 'N/A'), 's')
                santiago_now = datetime.now(ZoneInfo('UTC')).astimezone(ZoneInfo('America/Santiago'))
                self._update_cell_preserve_style(summary_sheet, 'B5', santiago_now.strftime('%d/%m/%Y %H:%M'), 's')
                
                # Extraer summary del reporte
                summary = current_report.get('summary', {})
                total_alarms = summary.get('totalAlarms', 0)
                alarm_types = summary.get('alarmTypes', {})
                
                # Actualizar métricas
                self._update_cell_preserve_style(summary_sheet, 'B9', total_alarms, 'n')
                self._update_cell_preserve_style(summary_sheet, 'B10', len(alarm_types), 'n')
                self._update_cell_preserve_style(summary_sheet, 'B11', len(filtered_events), 'n')
                
                # Actualizar tabla de resumen por alarma
                alarm_summary_data = []
                for alarm_type, count in alarm_types.items():
                    alarm_name = self.alarm_name_mapping.get(alarm_type.lower(), alarm_type)
                    alarm_summary_data.append([alarm_name, count])
                
                # Insertar datos comenzando en A15, preservando estilos
                for i, (alarm_name, count) in enumerate(alarm_summary_data):
                    row = i + 15  # Comenzar en fila 15
                    self._update_cell_preserve_style(summary_sheet, f'A{row}', alarm_name, 's')
                    self._update_cell_preserve_style(summary_sheet, f'B{row}', count, 'n')
            
            # --- ACTUALIZAR HOJA DE EVENTOS FILTRADOS ---
            events_sheet = workbook.get_sheet_by_name('Eventos Filtrados')
            if events_sheet:
                # CORREGIDO: Obtener los estilos de las celdas de la primera fila de datos (fila 2)
                template_styles = {}
                if events_sheet.max_row >= 2:  # Si hay al menos una fila de datos
                    for col in range(1, 8):  # 7 columnas (A-G)
                        col_letter = get_column_letter(col)
                        template_cell = events_sheet[f"{col_letter}2"]  # Primera fila de datos
                        template_styles[col_letter] = self._get_cell_style(template_cell)
                else:
                    # Si no hay filas de datos, obtener estilos de los encabezados (fila 1)
                    for col in range(1, 8):  # 7 columnas (A-G)
                        col_letter = get_column_letter(col)
                        template_cell = events_sheet[f"{col_letter}1"]  # Encabezados
                        template_styles[col_letter] = self._get_cell_style(template_cell)
                
                # Limpiar solo los valores existentes en filas de datos (desde fila 2)
                # Mantener exactamente la misma estructura de la plantilla
                if events_sheet.max_row > 1:
                    for row in range(2, events_sheet.max_row + 1):
                        for col in range(1, 8):  # 7 columnas (A-G)
                            cell_ref = f"{get_column_letter(col)}{row}"
                            if cell_ref in events_sheet:
                                # Limpiar solo el valor, mantener el estilo intacto
                                events_sheet[cell_ref].value = None
                
                # CORREGIDO: Insertar nuevos datos usando los estilos de la plantilla
                # Agregar filas si es necesario, pero manteniendo el formato
                events_processed = 0
                for i, event in enumerate(filtered_events):
                    row = i + 2  # Comenzar en fila 2 (después de encabezados)
                    
                    # Si necesitamos más filas de las disponibles, agregar nueva fila
                    if row > events_sheet.max_row:
                        # CORREGIDO: Usar el método correcto para agregar una fila
                        # Insertar una nueva fila al final con celdas vacías pero con estilos
                        new_row = []
                        for col in range(1, 8):  # 7 columnas (A-G)
                            col_letter = get_column_letter(col)
                            new_row.append('')  # Valor vacío
                        
                        # Insertar la nueva fila
                        events_sheet.append(new_row)
                        row = events_sheet.max_row  # Actualizar el número de fila
                        
                        # Aplicar estilos a las nuevas celdas
                        for col in range(1, 8):  # 7 columnas (A-G)
                            col_letter = get_column_letter(col)
                            new_cell = events_sheet[f"{col_letter}{row}"]
                            
                            # CORREGIDO: Aplicar estilos de plantilla pero con color de fondo específico para filas de datos
                            if col_letter in template_styles:
                                # Copiar el estilo de la plantilla
                                self._apply_cell_style(new_cell, template_styles[col_letter])
                                
                                # CORREGIDO: Para filas de datos (desde fila 2 hacia abajo), aplicar color de fondo #FDE9D9
                                # Solo si no es la fila de encabezados
                                if row >= 2:
                                    new_cell.fill = PatternFill(
                                        fill_type="solid",
                                        start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                        end_color="FFFDE9D9"
                                    )
                            else:
                                # Si no hay estilos de plantilla, crear estilo básico con color de fondo
                                if row >= 2:
                                    new_cell.fill = PatternFill(
                                        fill_type="solid",
                                        start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                        end_color="FFFDE9D9"
                                    )
                                else:
                                    # Para encabezados, usar estilo por defecto
                                    new_cell.font = Font(name='Arial', size=10, bold=True)
                    
                    # Enumeración (Columna A) - usar estilo de plantilla
                    if 'A' in template_styles:
                        cell = events_sheet[f'A{row}']
                        cell.value = i + 1
                        cell.data_type = 'n'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['A'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'A{row}', i + 1, 'n')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'A{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Fecha y Hora (Columna B) - usar estilo de plantilla
                    formatted_date = self._format_timestamp(event.get('timestamp', ''))
                    if 'B' in template_styles:
                        cell = events_sheet[f'B{row}']
                        cell.value = formatted_date
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['B'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'B{row}', formatted_date, 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'B{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Patente (Columna C) - usar estilo de plantilla
                    if 'C' in template_styles:
                        cell = events_sheet[f'C{row}']
                        cell.value = event.get('vehiclePlate', '')
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['C'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'C{row}', event.get('vehiclePlate', ''), 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'C{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Tipo de Alarma (Columna D) - usar estilo de plantilla
                    alarm_type = event.get('alarmType', '')
                    alarm_name = self.alarm_name_mapping.get(alarm_type.lower(), alarm_type)
                    if 'D' in template_styles:
                        cell = events_sheet[f'D{row}']
                        cell.value = alarm_name
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['D'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'D{row}', alarm_name, 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'D{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Conductor (Columna E) - usar estilo de plantilla
                    driver = event.get('driver', 'Sin conductor')
                    if 'E' in template_styles:
                        cell = events_sheet[f'E{row}']
                        cell.value = driver
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['E'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'E{row}', driver, 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'E{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Empresa (Columna F) - usar estilo de plantilla
                    company = event.get('company', selected_company)
                    if 'F' in template_styles:
                        cell = events_sheet[f'F{row}']
                        cell.value = company
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['F'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'F{row}', company, 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'F{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    # Comentarios (Columna G) - usar estilo de plantilla
                    comments = event.get('comments', 'Sin comentarios')
                    if 'G' in template_styles:
                        cell = events_sheet[f'G{row}']
                        cell.value = comments
                        cell.data_type = 's'
                        # Copiar estilo de plantilla
                        self._apply_cell_style(cell, template_styles['G'])
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            cell.fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    else:
                        self._update_cell_preserve_style(events_sheet, f'G{row}', comments, 's')
                        # CORREGIDO: Aplicar color de fondo para filas de datos
                        if row >= 2:
                            events_sheet[f'G{row}'].fill = PatternFill(
                                fill_type="solid",
                                start_color="FFFDE9D9",  # #FDE9D9 en formato RGB
                                end_color="FFFDE9D9"
                            )
                    
                    events_processed += 1
                
                print(f"Eventos procesados en hoja de eventos: {events_processed}")
            
            # Guardar el workbook en un objeto bytes
            from io import BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.read()
            
        except Exception as e:
            print(f"Error al exportar Excel: {e}")
            raise HTTPException(
                status_code=500,
                detail=f"Error al generar archivo Excel: {str(e)}"
            )
        finally:
            # Limpiar archivos temporales
            if template_path:
                self._cleanup_temp_file(template_path)

# Instancia global del servicio
excel_export_service_v2 = ExcelExportServiceV2()
